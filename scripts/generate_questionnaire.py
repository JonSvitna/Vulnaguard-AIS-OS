#!/usr/bin/env python3
"""
Generate a Vulnaguard discovery questionnaire as a formatted .xlsx file.

Supports all service codes:
    VA, CMMC-L2, NIST-171, SPA, ERA, POLICY, SENTINEL

Usage:
    python3 scripts/generate_questionnaire.py \
        --client "AfterSwing" \
        --service-code "NIST-171" \
        --service-name "NIST SP 800-171 Gap Assessment" \
        --due "2026-07-08" \
        --output "playbook/engagements/.../AfterSwing_DiscoveryQuestionnaire.xlsx"
"""
import argparse
from pathlib import Path
from datetime import date
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Brand colors ────────────────────────────────────────────────────────────
NAVY        = "0A1628"
BLUE        = "1B4FD8"
LIGHT_BLUE  = "EEF2FF"
WHITE       = "FFFFFF"
GRAY_BG     = "F8F9FA"
GRAY_BORDER = "D1D5DB"
YELLOW      = "FEF3C7"
DISABLED    = "EEEEEE"
DISABLED_TXT= "BBBBBB"


def thin(color=GRAY_BORDER):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)


# ── Universal base questions (all services) ──────────────────────────────────
BASE = [
    ("1. Organization Overview",
     "What is your organization's primary business and industry?", "text"),
    ("1. Organization Overview",
     "How many employees does your organization have? How many are involved with IT or security?", "text"),
    ("1. Organization Overview",
     "Where are your offices or facilities located?", "text"),
    ("1. Organization Overview",
     "Do you have remote or hybrid workers? Approximately how many?", "text"),
    ("1. Organization Overview",
     "Does your organization hold any federal government contracts? If yes, list agencies and whether contracts contain DFARS 252.204-7012 or CUI-related clauses.", "text"),

    ("2. IT Environment",
     "Describe your IT environment at a high level (on-premises, cloud, SaaS, hybrid).", "text"),
    ("2. IT Environment",
     "List the primary operating systems in use (Windows, macOS, Linux — versions if known).", "text"),
    ("2. IT Environment",
     "What cloud platforms or services do you use? (AWS, Azure, M365, Google Workspace, etc.)", "text"),
    ("2. IT Environment",
     "How many endpoints (laptops, desktops, workstations) are in scope?", "text"),
    ("2. IT Environment",
     "How many servers are in scope (physical or virtual)?", "text"),
    ("2. IT Environment",
     "Do you have a current network diagram?", "yesno"),
    ("2. IT Environment",
     "Do you use a VPN? If yes, what product?", "yesno"),
    ("2. IT Environment",
     "Do you have an MSSP or outsourced IT provider? If yes, who and what do they manage?", "yesno"),

    ("3. Security Controls",
     "Do you have a dedicated security role (CISO, IT Security Manager, etc.)?", "yesno"),
    ("3. Security Controls",
     "Do you have a documented Information Security Policy? When was it last reviewed?", "yesno"),
    ("3. Security Controls",
     "Do you require MFA for remote access (VPN)?", "yesno"),
    ("3. Security Controls",
     "Do you require MFA for email (M365 / Google)?", "yesno"),
    ("3. Security Controls",
     "Do you require MFA for critical systems and admin accounts?", "yesno"),
    ("3. Security Controls",
     "Do you have a formal patch management process? How frequently are patches applied?", "yesno"),
    ("3. Security Controls",
     "Do you use an endpoint detection and response (EDR) tool? If yes, which product?", "yesno"),
    ("3. Security Controls",
     "Do you have centralized logging and/or a SIEM? If yes, which product?", "yesno"),
    ("3. Security Controls",
     "Do you perform regular data backups? What is your backup frequency and retention? Are backups tested?", "yesno"),
    ("3. Security Controls",
     "Do you have an incident response plan? Has it been tested in the last 12 months?", "yesno"),
    ("3. Security Controls",
     "Do you conduct security awareness training? How frequently?", "yesno"),

    ("4. Stakeholders",
     "Who is the primary technical contact for this engagement? (Name, title, email, phone)", "text"),
    ("4. Stakeholders",
     "Who is the business/executive sponsor? (Name, title, email)", "text"),
    ("4. Stakeholders",
     "Who should receive status updates and reports? (Names and emails)", "text"),
    ("4. Stakeholders",
     "Who is the emergency contact during the engagement? (Name, phone — reachable outside business hours)", "text"),

    ("5. Additional Context",
     "Are there any systems, time windows, or activities to avoid during the engagement?", "text"),
    ("5. Additional Context",
     "Are there any regulatory, contractual, or other constraints we should be aware of?", "text"),
    ("5. Additional Context",
     "Anything else you'd like us to know about your environment or goals for this engagement?", "text"),
]

# ── Service-specific question sections ──────────────────────────────────────
SERVICE_SECTIONS = {

    "VA": [
        ("VA — Scope & Authorization",
         "List all IP addresses, IP ranges, and hostnames authorized for scanning. Be specific — we can only test what is listed here.", "text"),
        ("VA — Scope & Authorization",
         "List any IP addresses, ranges, or systems that must NOT be scanned under any circumstances.", "text"),
        ("VA — Scope & Authorization",
         "Are any web applications in scope? If yes, list URLs and whether authenticated testing is required.", "yesno"),
        ("VA — Scope & Authorization",
         "Are there preferred testing windows (days/hours) to reduce business impact?", "text"),
        ("VA — Scope & Authorization",
         "Does your network require IP whitelisting before our scanner can reach targets?", "yesno"),
        ("VA — Scope & Authorization",
         "Will VPN access be provided for internal scanning? If yes, who provisions this?", "yesno"),
        ("VA — Scope & Authorization",
         "For web app testing: can you create dedicated test accounts (authenticated + unauthenticated)? Who provisions them?", "yesno"),
        ("VA — Prior Scan History",
         "Have you run a vulnerability scan in the past 12 months? If yes, share results if available.", "yesno"),
        ("VA — Prior Scan History",
         "Are there known critical or high vulnerabilities you are already aware of?", "yesno"),
        ("VA — Prior Scan History",
         "Do you have an existing patch management schedule we should coordinate with?", "yesno"),
    ],

    "CMMC-L2": [
        ("CMMC — CUI Scoping",
         "What types of Controlled Unclassified Information (CUI) does your organization handle? (e.g., export controlled, technical data, procurement info)", "text"),
        ("CMMC — CUI Scoping",
         "Where is CUI created, stored, processed, or transmitted? List all systems, applications, and cloud services.", "text"),
        ("CMMC — CUI Scoping",
         "Who has access to CUI? (Roles — not individual names)", "text"),
        ("CMMC — CUI Scoping",
         "Do any third-party vendors or subcontractors have access to CUI? If yes, list them.", "yesno"),
        ("CMMC — CUI Scoping",
         "Have you defined a formal system boundary for your CUI environment?", "yesno"),
        ("CMMC — Documentation",
         "Do you have an existing System Security Plan (SSP)?", "yesno"),
        ("CMMC — Documentation",
         "Have you conducted a prior self-assessment against NIST SP 800-171?", "yesno"),
        ("CMMC — Documentation",
         "Do you have an existing POA&M for open controls?", "yesno"),
        ("CMMC — Documentation",
         "Have you submitted a SPRS score? If yes, what is your current score?", "yesno"),
        ("CMMC — Documentation",
         "List all DoD contracts. Confirm which contain the DFARS 252.204-7012 clause.", "text"),
        ("CMMC — Documentation",
         "Do you have a current subcontractor list with their level of CUI access documented?", "yesno"),
        ("CMMC — Compliance History",
         "Have you previously engaged a C3PAO or CMMC consultant? What was the outcome?", "yesno"),
        ("CMMC — Compliance History",
         "Do you have a target date for your formal C3PAO assessment?", "text"),
    ],

    "NIST-171": [
        ("NIST 800-171 — CUI Scope",
         "What types of Controlled Unclassified Information (CUI) does your organization handle?", "text"),
        ("NIST 800-171 — CUI Scope",
         "Where is CUI created, stored, processed, or transmitted? (List systems, apps, cloud services.)", "text"),
        ("NIST 800-171 — CUI Scope",
         "Who has access to CUI? (Roles — not individual names)", "text"),
        ("NIST 800-171 — CUI Scope",
         "Do any third-party vendors or subcontractors have access to CUI? If yes, list them.", "yesno"),
        ("NIST 800-171 — Documentation",
         "Do you have an existing System Security Plan (SSP) or draft?", "yesno"),
        ("NIST 800-171 — Documentation",
         "Have you conducted a prior self-assessment against NIST SP 800-171?", "yesno"),
        ("NIST 800-171 — Documentation",
         "Do you have an existing POA&M for open controls?", "yesno"),
        ("NIST 800-171 — Documentation",
         "Have you calculated or submitted a SPRS score? If yes, what is it?", "yesno"),
        ("NIST 800-171 — Documentation",
         "Do you have existing security policies and procedures? Please list them.", "yesno"),
        ("NIST 800-171 — Compliance Context",
         "What is driving this assessment? (DoD contract, internal initiative, cyber insurance, other)", "text"),
        ("NIST 800-171 — Compliance Context",
         "Is CMMC certification a future goal? If yes, what is the target timeline?", "yesno"),
    ],

    "SPA": [
        ("Security Program — Governance",
         "Who owns the security function? (Name, title, reporting line)", "text"),
        ("Security Program — Governance",
         "Does your organization have a security committee or risk committee?", "yesno"),
        ("Security Program — Governance",
         "How frequently does leadership receive a security briefing?", "text"),
        ("Security Program — Governance",
         "What is your annual security budget (approximate range)?", "text"),
        ("Security Program — Risk & Compliance",
         "Do you have an active risk register?", "yesno"),
        ("Security Program — Risk & Compliance",
         "Do you carry cyber liability insurance? What are the coverage limits?", "yesno"),
        ("Security Program — Risk & Compliance",
         "Have you had a security incident in the past 24 months? Describe briefly.", "yesno"),
        ("Security Program — Risk & Compliance",
         "What regulatory frameworks apply to your organization? (HIPAA, PCI-DSS, DFARS, state laws, etc.)", "text"),
        ("Security Program — Vendor & Third-Party",
         "Do you have a formal vendor/third-party risk management process?", "yesno"),
        ("Security Program — Vendor & Third-Party",
         "How many third-party vendors have access to your systems or sensitive data?", "text"),
        ("Security Program — People & Training",
         "Do you conduct annual security awareness training for all employees?", "yesno"),
        ("Security Program — People & Training",
         "Do you run phishing simulations? How frequently?", "yesno"),
        ("Security Program — People & Training",
         "Do new employees receive security onboarding?", "yesno"),
        ("Security Program — Prior Assessments",
         "Have you had a prior security assessment, audit, or penetration test? If yes, share the most recent report if available.", "yesno"),
        ("Security Program — Prior Assessments",
         "Are there known open findings from prior assessments?", "yesno"),
    ],

    "ERA": [
        ("Executive Risk — Business Context",
         "What are your organization's top 3 business objectives for the next 12 months?", "text"),
        ("Executive Risk — Business Context",
         "What are the 3–5 most critical systems or processes that, if disrupted, would most impact revenue or operations?", "text"),
        ("Executive Risk — Business Context",
         "What does a 'worst day' look like for your organization? (ransomware, data breach, regulatory fine, reputational damage)", "text"),
        ("Executive Risk — Business Context",
         "Are you pursuing any M&A activity, fundraising, or major contracts in the next 12 months?", "yesno"),
        ("Executive Risk — Risk Awareness",
         "Have you experienced a security incident in the past 24 months? Describe briefly.", "yesno"),
        ("Executive Risk — Risk Awareness",
         "Have you received any regulatory notices or customer security questionnaires you struggled to answer?", "yesno"),
        ("Executive Risk — Risk Awareness",
         "Do you carry cyber liability insurance? What are the coverage limits and deductible?", "yesno"),
        ("Executive Risk — Risk Awareness",
         "What regulatory requirements apply to your organization? (HIPAA, PCI, DFARS, state breach laws, etc.)", "text"),
        ("Executive Risk — Board & Leadership",
         "How many board members have a technology or security background?", "text"),
        ("Executive Risk — Board & Leadership",
         "How frequently does the board receive a cybersecurity update?", "text"),
        ("Executive Risk — Board & Leadership",
         "Who on the executive team owns cybersecurity risk? (Name, title)", "text"),
        ("Executive Risk — Vendors",
         "List your top 5 vendors who have access to your systems or sensitive data.", "text"),
        ("Executive Risk — Vendors",
         "Have any of your vendors experienced a breach in the past 12 months?", "yesno"),
    ],

    "POLICY": [
        ("Policy — Existing Documentation",
         "List all security policies your organization currently has in place (even informal or draft ones).", "text"),
        ("Policy — Existing Documentation",
         "Do you have an employee handbook with security or acceptable use provisions?", "yesno"),
        ("Policy — Existing Documentation",
         "When were your existing policies last reviewed or updated?", "text"),
        ("Policy — Existing Documentation",
         "Do your policies align to a specific framework (NIST 800-171, ISO 27001, CIS Controls)?", "yesno"),
        ("Policy — Regulatory Drivers",
         "What regulatory or contractual requirements are driving this policy work? (CMMC, HIPAA, PCI, state law, customer contract)", "text"),
        ("Policy — Regulatory Drivers",
         "Do you have a target compliance date or audit deadline?", "text"),
        ("Policy — Regulatory Drivers",
         "Has a prior assessment identified specific policy gaps? If yes, share the report if available.", "yesno"),
        ("Policy — Approval & Governance",
         "Who has authority to approve and sign policies? (Name, title)", "text"),
        ("Policy — Approval & Governance",
         "Who will own ongoing policy maintenance after delivery? (Name, title)", "text"),
        ("Policy — Approval & Governance",
         "What is your internal policy review cycle target? (Annual, bi-annual, etc.)", "text"),
        ("Policy — Scope",
         "Which of the following policies do you need developed or updated? (Mark all that apply in Comments: Information Security, Access Control, Incident Response, Patch Management, Change Management, Media Protection, Physical Security, Remote Work, Acceptable Use, Data Classification, Vendor Risk, Business Continuity, other)", "text"),
    ],

    "SENTINEL": [
        ("Sentinel — Subscription",
         "Which Sentinel subscription tier are you onboarding to?", "text"),
        ("Sentinel — Subscription",
         "What compliance framework will you track in Sentinel? (CMMC L2, NIST 800-171, other)", "text"),
        ("Sentinel — Users",
         "List all users to be provisioned. For each: Name, Email, Role (Admin / Contributor / Read-Only).", "text"),
        ("Sentinel — Users",
         "Who is the primary admin contact for your Sentinel tenant? (Name, email)", "text"),
        ("Sentinel — Prior Assessment Artifacts",
         "Do you have an existing POA&M from a prior assessment to import?", "yesno"),
        ("Sentinel — Prior Assessment Artifacts",
         "Do you have an existing evidence library (policies, configs, screenshots) to upload?", "yesno"),
        ("Sentinel — Prior Assessment Artifacts",
         "Do you have a defined CUI system boundary or asset inventory to import?", "yesno"),
        ("Sentinel — Environment",
         "List all IT systems and cloud services to be tracked in Sentinel.", "text"),
        ("Sentinel — Environment",
         "Do you have an existing asset inventory document? If yes, share it.", "yesno"),
        ("Sentinel — Reporting",
         "Who should receive automated compliance status reports? (Names and emails)", "text"),
        ("Sentinel — Reporting",
         "How frequently would you like automated reports delivered? (Weekly / Monthly / Quarterly)", "text"),
        ("Sentinel — Training",
         "How many staff will attend the Sentinel training session?", "text"),
        ("Sentinel — Training",
         "Do you prefer a live walkthrough session or recorded training?", "text"),
    ],
}


def get_questions(service_code: str) -> list:
    service_qs = SERVICE_SECTIONS.get(service_code, [])
    return BASE + service_qs


def build(client: str, service_code: str, service_name: str, due: str, output: Path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Discovery Questionnaire"

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 55
    ws.column_dimensions["C"].width = 9
    ws.column_dimensions["D"].width = 9
    ws.column_dimensions["E"].width = 9
    ws.column_dimensions["F"].width = 42

    # ── Header ───────────────────────────────────────────────────────
    ws.merge_cells("A1:F1")
    h1 = ws["A1"]
    h1.value = "VULNAGUARD LLC"
    h1.font = Font(name="Calibri", bold=True, size=16, color=WHITE)
    h1.fill = PatternFill("solid", fgColor=NAVY)
    h1.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:F2")
    h2 = ws["A2"]
    h2.value = "DISCOVERY QUESTIONNAIRE"
    h2.font = Font(name="Calibri", bold=True, size=13, color=WHITE)
    h2.fill = PatternFill("solid", fgColor=BLUE)
    h2.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[2].height = 24

    meta = [
        ("CLIENT",      client),
        ("ENGAGEMENT",  service_name),
        ("SERVICE CODE",service_code),
        ("DUE DATE",    due),
        ("PREPARED BY", "Vulnaguard LLC  |  seanmurrill@vulnaguard.com  |  vulnaguard.com"),
    ]
    for i, (label, value) in enumerate(meta, start=3):
        ws.merge_cells(f"A{i}:B{i}")
        ws.merge_cells(f"C{i}:F{i}")
        lc = ws[f"A{i}"]
        lc.value = f"  {label}"
        lc.font = Font(name="Calibri", bold=True, size=10, color=NAVY)
        lc.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
        lc.alignment = Alignment(vertical="center")
        lc.border = thin()
        vc = ws[f"C{i}"]
        vc.value = f"  {value}"
        vc.font = Font(name="Calibri", size=10)
        vc.alignment = Alignment(vertical="center", wrap_text=True)
        vc.border = thin()
        ws.row_dimensions[i].height = 18

    ins_row = len(meta) + 3
    ws.merge_cells(f"A{ins_row}:F{ins_row}")
    ins = ws[f"A{ins_row}"]
    ins.value = (
        "Instructions: For Yes/No/N/A questions, type X in the appropriate column. "
        "Use the Comments column for context, caveats, or open-text answers. "
        "Greyed Yes/No/N/A cells indicate open-text questions — use Comments only. "
        f"Return the completed file to seanmurrill@vulnaguard.com by {due}."
    )
    ins.font = Font(name="Calibri", italic=True, size=9, color="555555")
    ins.fill = PatternFill("solid", fgColor=YELLOW)
    ins.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
    ins.border = thin()
    ws.row_dimensions[ins_row].height = 40

    # Column headers
    col_row = ins_row + 2
    for col_idx, hdr in enumerate(["#", "Question", "Yes", "No", "N/A", "Comments / Notes"], start=1):
        cell = ws.cell(row=col_row, column=col_idx, value=hdr)
        cell.font = Font(name="Calibri", bold=True, size=10, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin(WHITE)
    ws.row_dimensions[col_row].height = 20

    # ── Questions ────────────────────────────────────────────────────
    questions = get_questions(service_code)
    current_section = None
    q_num = 1
    data_row = col_row + 1

    for (section, question, answer_type) in questions:
        if section != current_section:
            current_section = section
            ws.merge_cells(f"A{data_row}:F{data_row}")
            sc = ws[f"A{data_row}"]
            sc.value = f"  {section.upper()}"
            sc.font = Font(name="Calibri", bold=True, size=10, color=WHITE)
            sc.fill = PatternFill("solid", fgColor=BLUE)
            sc.alignment = Alignment(vertical="center")
            sc.border = thin(BLUE)
            ws.row_dimensions[data_row].height = 20
            data_row += 1

        row_fill = PatternFill("solid", fgColor=GRAY_BG if q_num % 2 == 0 else WHITE)

        num_cell = ws.cell(row=data_row, column=1, value=q_num)
        num_cell.font = Font(name="Calibri", size=9, color="888888")
        num_cell.alignment = Alignment(horizontal="center", vertical="top")
        num_cell.fill = row_fill
        num_cell.border = thin()

        q_cell = ws.cell(row=data_row, column=2, value=question)
        q_cell.font = Font(name="Calibri", size=10)
        q_cell.alignment = Alignment(vertical="top", wrap_text=True)
        q_cell.fill = row_fill
        q_cell.border = thin()

        for col_idx in range(3, 6):
            c = ws.cell(row=data_row, column=col_idx)
            c.alignment = Alignment(horizontal="center", vertical="top")
            if answer_type == "yesno":
                c.fill = row_fill
            else:
                c.fill = PatternFill("solid", fgColor=DISABLED)
                c.value = "—"
                c.font = Font(name="Calibri", size=9, color=DISABLED_TXT)
            c.border = thin()

        cmt = ws.cell(row=data_row, column=6)
        cmt.alignment = Alignment(vertical="top", wrap_text=True)
        cmt.fill = row_fill
        cmt.border = thin()

        ws.row_dimensions[data_row].height = 38
        q_num += 1
        data_row += 1

    # ── Footer ───────────────────────────────────────────────────────
    data_row += 1
    ws.merge_cells(f"A{data_row}:F{data_row}")
    ft = ws[f"A{data_row}"]
    ft.value = (
        "Vulnaguard LLC  |  vulnaguard.com  |  seanmurrill@vulnaguard.com  "
        "|  CAGE: 21QQ7  |  UEI: CJ7ZXTUSN3W8  |  Confidential"
    )
    ft.font = Font(name="Calibri", size=8, italic=True, color="888888")
    ft.alignment = Alignment(horizontal="center", vertical="center")
    ft.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    ws.row_dimensions[data_row].height = 16

    ws.freeze_panes = f"A{col_row + 1}"

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    print(f"Created: {output}  ({q_num - 1} questions)")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--client",       required=True)
    parser.add_argument("--service-code", required=True,
                        choices=["VA","CMMC-L2","NIST-171","SPA","ERA","POLICY","SENTINEL"])
    parser.add_argument("--service-name", required=True)
    parser.add_argument("--due",          default=str(date.today()))
    parser.add_argument("--output",       required=True)
    args = parser.parse_args()
    build(args.client, args.service_code, args.service_name, args.due, Path(args.output))


if __name__ == "__main__":
    main()
