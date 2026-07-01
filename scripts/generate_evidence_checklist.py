#!/usr/bin/env python3
"""
Generate a Vulnaguard evidence checklist as a formatted .xlsx file.

Supports all service codes:
    VA, CMMC-L2, NIST-171, SPA, ERA, POLICY, SENTINEL

Each row has: # | Evidence Item | Status | Received Date | Location | Notes

Usage:
    python3 scripts/generate_evidence_checklist.py \
        --client "AfterSwing" \
        --service-code "NIST-171" \
        --service-name "NIST SP 800-171 Gap Assessment" \
        --output "playbook/engagements/.../AfterSwing_EvidenceChecklist.xlsx"
"""
import argparse
from pathlib import Path
from datetime import date
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Brand colors ────────────────────────────────────────────────────────────
NAVY        = "0A1628"
BLUE        = "1B4FD8"
LIGHT_BLUE  = "EEF2FF"
WHITE       = "FFFFFF"
GRAY_BG     = "F8F9FA"
GRAY_BORDER = "D1D5DB"
YELLOW      = "FEF3C7"

STATUS_OPTIONS = "Not Requested / Requested / Received / Reviewed / N/A"


def thin(color=GRAY_BORDER):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)


# ── Universal evidence items (all services) ─────────────────────────────────
BASE = [
    ("Organizational",  "Organizational chart showing IT and security roles"),
    ("Organizational",  "List of primary points of contact (name, title, email, phone)"),
    ("Organizational",  "List of all physical facility locations in scope"),
    ("Organizational",  "Description of any third-party IT or MSSP relationships"),
    ("IT Environment",  "Current network diagram (logical and/or physical)"),
    ("IT Environment",  "Asset inventory — all endpoints, servers, network devices"),
    ("IT Environment",  "Cloud service and SaaS application inventory"),
    ("IT Environment",  "Current software/application inventory for in-scope systems"),
    ("IT Environment",  "List of external-facing IP addresses and hostnames"),
    ("Security Policy", "Information Security Policy"),
    ("Security Policy", "Acceptable Use Policy"),
    ("Security Policy", "Access Control Policy"),
    ("Security Policy", "Incident Response Plan"),
    ("Security Policy", "Business Continuity / Disaster Recovery Plan"),
    ("Security Policy", "Data Classification Policy"),
    ("Access Control",  "Active Directory / identity provider configuration overview"),
    ("Access Control",  "Privileged access list (admin accounts and their access levels)"),
    ("Access Control",  "MFA enrollment report"),
    ("Access Control",  "Remote access configuration (VPN settings, approved methods)"),
    ("Logging",         "SIEM / logging platform overview or sample log exports"),
    ("Logging",         "Audit logging policy or configuration documentation"),
    ("Patch Mgmt",      "Patch management policy or procedure"),
    ("Patch Mgmt",      "Most recent vulnerability scan report (if available)"),
    ("Training",        "Security awareness training records (completion rates, dates)"),
    ("Training",        "Evidence of phishing simulation results (if conducted)"),
]

# ── Service-specific evidence items ─────────────────────────────────────────
SERVICE_SECTIONS = {

    "VA": [
        ("VA — Scope",          "Signed Rules of Engagement (ROE) / authorization letter"),
        ("VA — Scope",          "Authorized IP ranges, hostnames, and URLs list"),
        ("VA — Scope",          "List of explicitly excluded systems or addresses"),
        ("VA — Scope",          "Change management freeze calendar / blackout dates"),
        ("VA — Access",         "VPN credentials / access instructions for internal scanning"),
        ("VA — Access",         "Web application test accounts (authenticated and guest)"),
        ("VA — Prior Work",     "Most recent prior vulnerability scan report"),
        ("VA — Prior Work",     "Most recent penetration test report"),
        ("VA — Prior Work",     "Existing known vulnerability backlog or remediation tracker"),
    ],

    "CMMC-L2": [
        ("CMMC — CUI Scope",    "System Security Plan (SSP) — current version"),
        ("CMMC — CUI Scope",    "System boundary diagram (visual CUI environment boundary)"),
        ("CMMC — CUI Scope",    "CUI inventory — where CUI is created, stored, processed, transmitted"),
        ("CMMC — CUI Scope",    "List of systems and cloud services within the CUI boundary"),
        ("CMMC — CUI Scope",    "Subcontractor list with notation of which have CUI access"),
        ("CMMC — CUI Scope",    "All applicable DoD contracts (with DFARS 252.204-7012 clause highlighted)"),
        ("CMMC — Assessment",   "Prior self-assessment results or scoring workbook"),
        ("CMMC — Assessment",   "Current SPRS score submission confirmation (if submitted)"),
        ("CMMC — Assessment",   "Existing Plan of Action & Milestones (POA&M)"),
        ("CMMC — Policies",     "Media Protection Policy and procedures"),
        ("CMMC — Policies",     "Configuration Management Policy and baseline configs"),
        ("CMMC — Policies",     "System and Communications Protection documentation"),
        ("CMMC — Policies",     "Risk Assessment documentation"),
        ("CMMC — Policies",     "Personnel Security Policy"),
    ],

    "NIST-171": [
        ("NIST 800-171 — CUI",  "System Security Plan (SSP) — current version or draft"),
        ("NIST 800-171 — CUI",  "System boundary diagram / network diagram with CUI flows"),
        ("NIST 800-171 — CUI",  "CUI inventory or data flow diagram"),
        ("NIST 800-171 — CUI",  "List of all systems and cloud services within the CUI boundary"),
        ("NIST 800-171 — CUI",  "Subcontractor list with notation of which have CUI access"),
        ("NIST 800-171 — CUI",  "Relevant DoD contracts with DFARS 252.204-7012 clause"),
        ("NIST 800-171 — Docs", "Prior self-assessment results or scoring workbook"),
        ("NIST 800-171 — Docs", "Existing POA&M (if any)"),
        ("NIST 800-171 — Docs", "SPRS score submission (if submitted)"),
        ("NIST 800-171 — Docs", "All existing security policies and procedures"),
        ("NIST 800-171 — Docs", "Configuration baselines for in-scope systems"),
    ],

    "SPA": [
        ("SPA — Governance",    "Risk management policy / framework documentation"),
        ("SPA — Governance",    "Active risk register"),
        ("SPA — Governance",    "Cyber liability insurance certificate / coverage summary"),
        ("SPA — Governance",    "Security committee meeting minutes (most recent 2 meetings)"),
        ("SPA — Incidents",     "Incident log / history for the past 24 months"),
        ("SPA — Incidents",     "Most recent incident response tabletop exercise results"),
        ("SPA — Vendor Risk",   "Vendor / third-party inventory"),
        ("SPA — Vendor Risk",   "Third-party risk assessment questionnaires or contracts (sample)"),
        ("SPA — People",        "Security awareness training program description"),
        ("SPA — People",        "Phishing simulation results (most recent)"),
        ("SPA — Prior Audits",  "Most recent external security assessment or audit report"),
        ("SPA — Prior Audits",  "Open findings tracker from prior assessments"),
    ],

    "ERA": [
        ("ERA — Business",      "Annual report or business overview (if publicly available)"),
        ("ERA — Business",      "List of 3–5 most critical business systems or processes"),
        ("ERA — Business",      "Cyber liability insurance certificate and coverage limits"),
        ("ERA — Incidents",     "Incident history for the past 24 months"),
        ("ERA — Incidents",     "Any regulatory notices, fines, or breach notifications"),
        ("ERA — Risk",          "Most recent external security assessment (if any)"),
        ("ERA — Risk",          "Active risk register (if one exists)"),
        ("ERA — Risk",          "Board or executive cybersecurity briefing materials (most recent)"),
        ("ERA — Regulatory",    "List of regulatory frameworks the organization is subject to"),
        ("ERA — Regulatory",    "Evidence of compliance activities (HIPAA, PCI, DFARS, etc.)"),
    ],

    "POLICY": [
        ("Policy — Existing",   "All existing security policies (even informal or draft versions)"),
        ("Policy — Existing",   "Employee handbook with any security / acceptable use provisions"),
        ("Policy — Existing",   "Policy review history (dates of last review per document)"),
        ("Policy — Regulatory", "Regulatory requirements driving policy development"),
        ("Policy — Regulatory", "Most recent gap assessment or audit noting policy deficiencies"),
        ("Policy — Process",    "Policy approval and distribution process documentation"),
        ("Policy — Process",    "Organizational chart showing policy ownership chain"),
        ("Policy — Process",    "HR onboarding / offboarding procedures referencing policies"),
    ],

    "SENTINEL": [
        ("Sentinel — Onboarding", "Sentinel subscription confirmation / tier details"),
        ("Sentinel — Onboarding", "User roster for provisioning (Name, Email, Role)"),
        ("Sentinel — Import",     "Existing POA&M to import (if available)"),
        ("Sentinel — Import",     "Existing evidence library or compliance artifact folder"),
        ("Sentinel — Import",     "Asset inventory for import into Sentinel asset tracker"),
        ("Sentinel — Import",     "System boundary diagram"),
        ("Sentinel — Config",     "List of compliance frameworks to activate in Sentinel"),
        ("Sentinel — Config",     "Report delivery preferences (recipients, frequency)"),
    ],
}


def get_items(service_code: str) -> list:
    service_items = SERVICE_SECTIONS.get(service_code, [])
    return BASE + service_items


def build(client: str, service_code: str, service_name: str, output: Path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Evidence Checklist"

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 52
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 28
    ws.column_dimensions["F"].width = 38

    # ── Header ────────────────────────────────────────────────────────
    ws.merge_cells("A1:F1")
    h1 = ws["A1"]
    h1.value = "VULNAGUARD LLC"
    h1.font = Font(name="Calibri", bold=True, size=16, color=WHITE)
    h1.fill = PatternFill("solid", fgColor=NAVY)
    h1.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:F2")
    h2 = ws["A2"]
    h2.value = "EVIDENCE CHECKLIST"
    h2.font = Font(name="Calibri", bold=True, size=13, color=WHITE)
    h2.fill = PatternFill("solid", fgColor=BLUE)
    h2.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[2].height = 24

    meta = [
        ("CLIENT",      client),
        ("ENGAGEMENT",  service_name),
        ("SERVICE CODE",service_code),
        ("DATE",        str(date.today())),
        ("PREPARED BY", "Vulnaguard LLC  |  seanmurrill@vulnaguard.com  |  vulnaguard.com"),
    ]
    for i, (label, value) in enumerate(meta, start=3):
        ws.merge_cells(f"A{i}:B{i}")
        ws.merge_cells(f"C{i}:F{i}")
        lc = ws[f"A{i}"]
        lc.value = f"  {label}"
        lc.font = Font(name="Calibri", bold=True, size=10, color=NAVY)
        lc.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
        lc.alignment = Alignment(vertical="center")
        lc.border = thin()
        vc = ws[f"C{i}"]
        vc.value = f"  {value}"
        vc.font = Font(name="Calibri", size=10)
        vc.alignment = Alignment(vertical="center")
        vc.border = thin()
        ws.row_dimensions[i].height = 18

    ins_row = len(meta) + 3
    ws.merge_cells(f"A{ins_row}:F{ins_row}")
    ins = ws[f"A{ins_row}"]
    ins.value = (
        f"Instructions: Track the status of each evidence item below. "
        f"Status options: {STATUS_OPTIONS}. "
        "Record the received date when documents arrive. Location should note where the file is stored "
        "(e.g., SharePoint path, email date, folder name). Use Notes for version info, gaps, or follow-up actions."
    )
    ins.font = Font(name="Calibri", italic=True, size=9, color="555555")
    ins.fill = PatternFill("solid", fgColor=YELLOW)
    ins.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
    ins.border = thin()
    ws.row_dimensions[ins_row].height = 40

    # Column headers
    col_row = ins_row + 2
    headers = ["#", "Evidence Item", "Status", "Received Date", "Location / File Path", "Notes"]
    for col_idx, hdr in enumerate(headers, start=1):
        cell = ws.cell(row=col_row, column=col_idx, value=hdr)
        cell.font = Font(name="Calibri", bold=True, size=10, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.alignment = Alignment(horizontal="center" if col_idx != 2 else "left", vertical="center")
        cell.border = thin(WHITE)
    ws.row_dimensions[col_row].height = 20

    # ── Evidence items ────────────────────────────────────────────────
    items = get_items(service_code)
    current_section = None
    item_num = 1
    data_row = col_row + 1

    for (section, item) in items:
        if section != current_section:
            current_section = section
            ws.merge_cells(f"A{data_row}:F{data_row}")
            sc = ws[f"A{data_row}"]
            sc.value = f"  {section.upper()}"
            sc.font = Font(name="Calibri", bold=True, size=10, color=WHITE)
            sc.fill = PatternFill("solid", fgColor=BLUE)
            sc.alignment = Alignment(vertical="center")
            sc.border = thin(BLUE)
            ws.row_dimensions[data_row].height = 20
            data_row += 1

        row_fill = PatternFill("solid", fgColor=GRAY_BG if item_num % 2 == 0 else WHITE)

        num_cell = ws.cell(row=data_row, column=1, value=item_num)
        num_cell.font = Font(name="Calibri", size=9, color="888888")
        num_cell.alignment = Alignment(horizontal="center", vertical="center")
        num_cell.fill = row_fill
        num_cell.border = thin()

        item_cell = ws.cell(row=data_row, column=2, value=item)
        item_cell.font = Font(name="Calibri", size=10)
        item_cell.alignment = Alignment(vertical="center", wrap_text=True)
        item_cell.fill = row_fill
        item_cell.border = thin()

        status_cell = ws.cell(row=data_row, column=3, value="Not Requested")
        status_cell.font = Font(name="Calibri", size=9)
        status_cell.alignment = Alignment(horizontal="center", vertical="center")
        status_cell.fill = row_fill
        status_cell.border = thin()

        for col_idx in [4, 5, 6]:
            c = ws.cell(row=data_row, column=col_idx)
            c.fill = row_fill
            c.border = thin()
            c.alignment = Alignment(vertical="center", wrap_text=True)

        ws.row_dimensions[data_row].height = 22
        item_num += 1
        data_row += 1

    # ── Footer ────────────────────────────────────────────────────────
    data_row += 1
    ws.merge_cells(f"A{data_row}:F{data_row}")
    ft = ws[f"A{data_row}"]
    ft.value = (
        "Vulnaguard LLC  |  vulnaguard.com  |  seanmurrill@vulnaguard.com  "
        "|  CAGE: 21QQ7  |  UEI: CJ7ZXTUSN3W8  |  Confidential"
    )
    ft.font = Font(name="Calibri", size=8, italic=True, color="888888")
    ft.alignment = Alignment(horizontal="center", vertical="center")
    ft.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    ws.row_dimensions[data_row].height = 16

    ws.freeze_panes = f"A{col_row + 1}"

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    print(f"Created: {output}  ({item_num - 1} items)")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--client",       required=True)
    parser.add_argument("--service-code", required=True,
                        choices=["VA","CMMC-L2","NIST-171","SPA","ERA","POLICY","SENTINEL"])
    parser.add_argument("--service-name", required=True)
    parser.add_argument("--output",       required=True)
    args = parser.parse_args()
    build(args.client, args.service_code, args.service_name, Path(args.output))


if __name__ == "__main__":
    main()
